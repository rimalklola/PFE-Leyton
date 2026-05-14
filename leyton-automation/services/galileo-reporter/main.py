import os
import sys
import time

_SERVICES_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _SERVICES_ROOT not in sys.path:
    sys.path.insert(0, _SERVICES_ROOT)

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, date
from mock_data import MOCK_GALILEO_MISSIONS, MOCK_SERVICE_STATES, SEVERITY
from shared.registry import log_run, get_last_run, was_run_this_month
from shared.logger import ServiceLogger

OUTPUT_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "output")

DARK_BLUE = "003366"
GREEN = "E2EFDA"
YELLOW = "FFF2CC"
RED = "FCE4D6"
GREY = "F2F2F2"
WHITE = "FFFFFF"

BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

SERVICE_REGISTRY_NAMES = {
    "folder_creator": "folder-creator",
    "timesheet_prefill": "timesheet-prefill",
    "belspo_extractor": "belspo-extractor",
    "handover_generator": "handover-generator",
}

EXPIRY_WARN_DAYS = 30


def _cell(ws, row, col, value, bg=WHITE, bold=False):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
    c.font = Font(bold=bold, color="000000" if bg != DARK_BLUE else WHITE)
    c.alignment = CENTER
    c.border = BORDER
    return c


def _has_registry_data(service_key):
    reg_name = SERVICE_REGISTRY_NAMES.get(service_key)
    return reg_name is not None and get_last_run(reg_name) is not None


def _check_folder(client_id):
    if _has_registry_data("folder_creator"):
        last = get_last_run("folder-creator")
        if last and last["status"] == "success":
            return "OK", GREEN
        return "FAIL", RED
    state = MOCK_SERVICE_STATES.get("folder_creator", {}).get(client_id, {})
    if state.get("created"):
        return "OK", GREEN
    return "FAIL", RED


def _check_timesheet(client_id):
    if _has_registry_data("timesheet_prefill"):
        ran = was_run_this_month("timesheet-prefill")
        return ("OK", GREEN) if ran else ("WARNING", YELLOW)
    state = MOCK_SERVICE_STATES.get("timesheet_prefill", {}).get(client_id, {})
    if state.get("processed"):
        return "OK", GREEN
    return "WARNING", YELLOW


def _check_belspo(client_id, mission):
    if mission.get("belspo_notification_id") is None:
        return "N/A", GREY
    if _has_registry_data("belspo_extractor"):
        last = get_last_run("belspo-extractor")
        if last and last["status"] == "success":
            return "OK", GREEN
        return "WARNING", YELLOW
    state = MOCK_SERVICE_STATES.get("belspo_extractor", {}).get(client_id, {})
    if state.get("extracted"):
        return "OK", GREEN
    return "WARNING", YELLOW


def _check_handover(client_id, handover_required):
    if not handover_required:
        return "N/A", GREY
    if _has_registry_data("handover_generator"):
        last = get_last_run("handover-generator")
        if last and last["status"] == "success":
            return "OK", GREEN
        return "FAIL", RED
    state = MOCK_SERVICE_STATES.get("handover_generator", {}).get(client_id, {})
    if state.get("generated"):
        return "OK", GREEN
    return "FAIL", RED


def _check_expiry(mission):
    expiry_str = mission.get("belspo_expiry_date")
    if expiry_str is None:
        return "N/A", GREY
    try:
        expiry = date.fromisoformat(expiry_str)
    except ValueError:
        return "UNKNOWN", GREY
    today = date.today()
    if expiry < today:
        return f"EXPIRED ({expiry_str})", RED
    days_left = (expiry - today).days
    if days_left <= EXPIRY_WARN_DAYS:
        return f"EXPIRING ({days_left}d)", YELLOW
    return f"OK ({expiry_str})", GREEN


def _overall_severity(checks):
    statuses = [c[0] for c in checks]
    if any("FAIL" in s or "EXPIRED" in s for s in statuses):
        return SEVERITY["CRITICAL"], RED
    if any("WARNING" in s or "EXPIRING" in s for s in statuses):
        return SEVERITY["WARNING"], YELLOW
    return SEVERITY["OK"], GREEN


def build_alerts(mission, folder, timesheet, belspo, handover, expiry):
    alerts = []
    name = mission["client_name"]
    if "FAIL" in folder[0]:
        alerts.append(f"Folder not created for {name}")
    if "WARNING" in timesheet[0]:
        alerts.append(f"No timesheet processed this month for {name}")
    if "WARNING" in belspo[0]:
        alerts.append(f"Belspo extraction stale or missing for {name}")
    if "FAIL" in handover[0]:
        alerts.append(f"Handover required but not generated for {name}")
    if "EXPIRED" in expiry[0]:
        alerts.append(f"Belspo notification EXPIRED for {name}")
    elif "EXPIRING" in expiry[0]:
        alerts.append(f"Belspo notification expiring soon for {name}")
    if mission.get("belspo_notification_id") is None and mission["status"] == "active":
        alerts.append(f"No Belspo notification filed for {name}")
    return "; ".join(alerts) if alerts else "None"


def export_to_excel(rows, filepath):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Alert Matrix"

    headers = [
        "Mission ID", "Client", "Consultant", "Stage", "Status",
        "Folder Created", "Timesheet This Month", "Belspo Extracted",
        "Handover Generated", "Belspo Expiry", "Overall", "Alerts",
    ]
    col_widths = [16, 22, 18, 14, 10, 16, 20, 18, 20, 22, 18, 60]

    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
        c.font = Font(bold=True, color=WHITE)
        c.alignment = CENTER
        c.border = BORDER

    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    for row_idx, r in enumerate(rows, 2):
        _cell(ws, row_idx, 1, r["mission_id"])
        _cell(ws, row_idx, 2, r["client_name"])
        _cell(ws, row_idx, 3, r["consultant"])
        _cell(ws, row_idx, 4, r["stage"])
        _cell(ws, row_idx, 5, r["status"])
        _cell(ws, row_idx, 6, r["folder"][0], r["folder"][1])
        _cell(ws, row_idx, 7, r["timesheet"][0], r["timesheet"][1])
        _cell(ws, row_idx, 8, r["belspo"][0], r["belspo"][1])
        _cell(ws, row_idx, 9, r["handover"][0], r["handover"][1])
        _cell(ws, row_idx, 10, r["expiry"][0], r["expiry"][1])
        _cell(ws, row_idx, 11, r["overall"][0], r["overall"][1], bold=True)
        _cell(ws, row_idx, 12, r["alerts"])
        ws.row_dimensions[row_idx].height = 30

    wb.save(filepath)


def run():
    log = ServiceLogger("galileo-reporter")
    start = time.time()
    log.info("Service started")

    os.makedirs(OUTPUT_PATH, exist_ok=True)

    rows = []
    try:
        for mission in MOCK_GALILEO_MISSIONS:
            client_id = mission["client_id"]
            folder = _check_folder(client_id)
            timesheet = _check_timesheet(client_id)
            belspo = _check_belspo(client_id, mission)
            handover = _check_handover(client_id, mission.get("handover_required", False))
            expiry = _check_expiry(mission)
            overall = _overall_severity([folder, timesheet, belspo, handover, expiry])
            alerts = build_alerts(mission, folder, timesheet, belspo, handover, expiry)

            row = {
                "mission_id": mission["mission_id"],
                "client_name": mission["client_name"],
                "consultant": mission["consultant"],
                "stage": mission["stage"],
                "status": mission["status"],
                "folder": folder,
                "timesheet": timesheet,
                "belspo": belspo,
                "handover": handover,
                "expiry": expiry,
                "overall": overall,
                "alerts": alerts,
            }
            rows.append(row)
            log.info("Mission evaluated", mission_id=mission["mission_id"],
                     client=mission["client_name"], severity=overall[0])

        filename = f"Galileo_Report_{datetime.now().strftime('%Y%m%d')}.xlsx"
        filepath = os.path.join(OUTPUT_PATH, filename)
        export_to_excel(rows, filepath)

        critical = sum(1 for r in rows if r["overall"][0] == SEVERITY["CRITICAL"])
        warnings = sum(1 for r in rows if r["overall"][0] == SEVERITY["WARNING"])
        duration_ms = int((time.time() - start) * 1000)
        log_run("galileo-reporter", status="success", output_file=filepath, duration_ms=duration_ms)
        log.info("Service completed", missions=len(rows), critical=critical,
                 warnings=warnings, output_file=filepath, duration_ms=duration_ms)

    except Exception as exc:
        duration_ms = int((time.time() - start) * 1000)
        log.error("Service failed", error=str(exc))
        log_run("galileo-reporter", status="failed", error_message=str(exc), duration_ms=duration_ms)
        raise

    return rows


if __name__ == "__main__":
    run()
