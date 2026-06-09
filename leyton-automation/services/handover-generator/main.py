import os
import sys
import time

_SERVICES_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _SERVICES_ROOT not in sys.path:
    sys.path.insert(0, _SERVICES_ROOT)

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
from mock_data import MOCK_HANDOVER_DATA
from shared.registry import log_run
from shared.logger import ServiceLogger

OUTPUT_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "output")

DARK_BLUE = "003366"
LIGHT_BLUE = "D9E1F2"
LIGHT_GREY = "F2F2F2"
WHITE = "FFFFFF"
GREEN = "E2EFDA"
RED = "FCE4D6"


def header_style(cell, bg=DARK_BLUE, fg=WHITE, bold=True):
    cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
    cell.font = Font(color=fg, bold=bold)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )


def section_style(cell, bg=LIGHT_BLUE):
    cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
    cell.font = Font(bold=True, color="003366")
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )


def value_style(cell, bg=WHITE):
    cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
    cell.font = Font(color="000000")
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )


def write_row(ws, row, label, value, label_bg=LIGHT_GREY):
    label_cell = ws.cell(row=row, column=1, value=label)
    section_style(label_cell, bg=label_bg)
    value_cell = ws.cell(row=row, column=2, value=value)
    value_style(value_cell)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    ws.row_dimensions[row].height = 30


def write_link_row(ws, row, label, url, label_bg=LIGHT_GREY):
    """Like write_row but makes the value a clickable hyperlink if it starts with http."""
    label_cell = ws.cell(row=row, column=1, value=label)
    section_style(label_cell, bg=label_bg)
    value_cell = ws.cell(row=row, column=2, value=url or "—")
    value_style(value_cell)
    if url and (url.startswith("http://") or url.startswith("https://")):
        value_cell.hyperlink = url
        value_cell.font = Font(color="0563C1", underline="single")
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    ws.row_dimensions[row].height = 30


def write_section_header(ws, row, title):
    cell = ws.cell(row=row, column=1, value=title)
    header_style(cell)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.row_dimensions[row].height = 25


# ---------------------------------------------------------------------------
# Field extraction — reverse-parse our own Excel output to pre-fill the form
# ---------------------------------------------------------------------------

_LABEL_TO_FIELD = {
    # Team
    "ho done with":                 "outgoing_consultant",
    "tax consultant":               "incoming_consultant",
    "sales":                        "sales_contact",
    # Company info
    "sector of activity":           "sector",
    "type of company":              "company_type",
    "account history":              "account_history",
    "belspo vat number":            "belspo_vat",
    "belspo password":              "belspo_password",
    "belspo notification":          "belspo_notification",
    "previous control":             "previous_control",
    # Contract
    "tax measure":                  "mission_type",
    "expected year of application": "mission_start",
    "remuneration system":          "remuneration_system",
    "gdpr":                         "gdpr",
    # Contacts
    "technical contact":            "technical_contact",
    "hr contact":                   "hr_contact",
    "social secretary":             "social_secretary",
    "key client contacts":          "key_contacts",
    # Folder links
    "sharepoint / web url":         "folder_link",
    "network drive path":           "network_path",
    # Previous mission
    "general comments":             "notes",
    "next step":                    "pending_tasks",
    "introduction email":           "introduction_email",
}


def extract_fields(filepath: str) -> dict:
    """
    Parse an existing handover Excel and return a flat dict mapping
    form field names → extracted values, ready to pre-fill the UI form.

    Supports:
    - Our own generated handover format (label | value two-column layout)
    - Any two-column Excel where column A is a label and column B is the value
    """
    result = {}
    ext = os.path.splitext(filepath)[1].lower()
    if ext not in (".xlsx", ".xls"):
        return result

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active

        for row in ws.iter_rows(values_only=True):
            if not any(row):
                continue

            a = str(row[0] or "").strip()
            b = str(row[1] or "").strip() if len(row) > 1 and row[1] is not None else ""

            # Title row  →  "HANDOVER FILE - CLIENT NAME"
            if a.upper().startswith("HANDOVER FILE"):
                parts = a.split("-", 1)
                if len(parts) == 2:
                    client = parts[1].strip().title()
                    if client:
                        result["client_name"] = client
                continue

            label = a.lower().strip()
            # Skip blank values and placeholder dashes
            if label in _LABEL_TO_FIELD and b and b not in ("—", "-", "None", ""):
                result[_LABEL_TO_FIELD[label]] = b

    except Exception:
        pass   # return whatever was collected so far

    return result


def generate_handover(data: dict) -> str:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Handover Sheet"

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20

    row = 1

    title_cell = ws.cell(row=row, column=1, value=f"HANDOVER FILE - {data['account_name'].upper()}")
    title_cell.font = Font(bold=True, size=14, color=WHITE)
    title_cell.fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.row_dimensions[row].height = 35
    row += 1

    date_cell = ws.cell(row=row, column=1,
                        value=f"Generated on: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    date_cell.font = Font(italic=True, color="666666")
    date_cell.alignment = Alignment(horizontal="right")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    row += 2

    # ── FOLDER LINKS ───────────────────────────────────────────────
    links = data.get("folder_links", {})
    if links.get("sharepoint") or links.get("network_path"):
        write_section_header(ws, row, "CLIENT FOLDER LINKS")
        row += 1
        write_link_row(ws, row, "SharePoint / Web URL", links.get("sharepoint", "—"))
        row += 1
        write_row(ws, row, "Network Drive Path", links.get("network_path", "—"))
        row += 2

    write_section_header(ws, row, "TEAM COMPOSITION")
    row += 1
    write_row(ws, row, "HO Done With", data["team"]["ho_done_with"])
    row += 1
    write_row(ws, row, "Tax Consultant", data["team"]["tax_consultant"])
    row += 1
    write_row(ws, row, "Sales", data["team"]["sales"])
    row += 2

    write_section_header(ws, row, "COMPANY INFORMATION")
    row += 1
    ci = data["company_info"]
    write_row(ws, row, "Sector of Activity", ci["sector"])
    row += 1
    write_row(ws, row, "Type of Company", ci["type"])
    row += 1
    write_row(ws, row, "Account History", ci["account_history"])
    row += 1
    write_row(ws, row, "Belspo VAT Number", ci["belspo_vat"])
    row += 1
    write_row(ws, row, "Belspo Password", ci["belspo_password"])
    row += 1
    write_row(ws, row, "Belspo Notification", ci["belspo_notification"])
    row += 1
    write_row(ws, row, "Previous Control", ci["previous_control"])
    row += 2

    write_section_header(ws, row, "CONTRACT INFORMATION")
    row += 1
    co = data["contract_info"]
    write_row(ws, row, "Tax Measure", co["tax_measure"])
    row += 1
    write_row(ws, row, "Expected Year of Application", co["expected_year"])
    row += 1
    write_row(ws, row, "Remuneration System", co["remuneration_system"])
    row += 1
    write_row(ws, row, "GDPR", co["gdpr"])
    row += 2

    write_section_header(ws, row, "POINTS OF CONTACT")
    row += 1
    poc = data["pocs"]
    write_row(ws, row, "Technical Contact",  poc.get("technical_contact", "—"))
    row += 1
    write_row(ws, row, "HR Contact",         poc.get("hr_contact", "—"))
    row += 1
    write_row(ws, row, "Social Secretary",   poc.get("social_secretary", "—"))
    row += 1
    write_row(ws, row, "Key Client Contacts", poc.get("key_contacts", "—"))
    row += 2

    write_section_header(ws, row, "PREVIOUS MISSION INFORMATION")
    row += 1
    pm = data["previous_mission"]
    write_row(ws, row, "Technical Report", pm["technical_report"])
    row += 1
    write_row(ws, row, "Timesheets", pm["timesheets"])
    row += 1
    write_row(ws, row, "Number of Projects", str(pm["nbr_projects"]))
    row += 1
    write_row(ws, row, "Number of Eligible Employees", str(pm["nbr_eligible_employees"]))
    row += 1
    write_row(ws, row, "Structural Research Certificate", pm["structural_research_certificate"])
    row += 1
    write_row(ws, row, "Mission Status", pm["mission_status"])
    row += 1
    write_row(ws, row, "Control", pm["control"])
    row += 1
    write_row(ws, row, "General Comments", pm["general_comments"])
    row += 1
    write_row(ws, row, "Next Step", pm["next_step"])
    row += 1
    write_row(ws, row, "Introduction Email", pm["introduction_email"])
    row += 2

    write_section_header(ws, row, "BELSPO NOTIFICATIONS")
    row += 1
    for col, h in enumerate(["Notification ID", "Project Title", "Status", "Covers Until"], 1):
        cell = ws.cell(row=row, column=col, value=h)
        header_style(cell, bg=LIGHT_BLUE, fg="003366")
    row += 1
    for notif in data["belspo_notifications"]:
        ws.cell(row=row, column=1, value=notif["notification_id"])
        ws.cell(row=row, column=2, value=notif["project_title"])
        ws.cell(row=row, column=3, value=notif["status"])
        ws.cell(row=row, column=4, value=notif["covers_until"])
        for col in range(1, 5):
            value_style(ws.cell(row=row, column=col))
        row += 1
    row += 1

    write_section_header(ws, row, "ELIGIBLE EMPLOYEES")
    row += 1
    for col, h in enumerate(["Employee Name", "Diploma", "R&D %", ""], 1):
        cell = ws.cell(row=row, column=col, value=h)
        header_style(cell, bg=LIGHT_BLUE, fg="003366")
    row += 1
    for emp in data["employees"]:
        ws.cell(row=row, column=1, value=emp["name"])
        ws.cell(row=row, column=2, value=emp["diploma"])
        ws.cell(row=row, column=3, value=f"{emp['rd_percentage']}%")
        for col in range(1, 4):
            value_style(ws.cell(row=row, column=col))
        row += 1
    row += 1

    write_section_header(ws, row, "DATA CHECKLIST")
    row += 1
    checklist = data["data_checklist"]
    items = {
        "Calculation Sheets": checklist["calculation_sheets"],
        "Control Documents": checklist["control_documents"],
        "Diplomas": checklist["diplomas"],
        "Individual Accounts": checklist["individual_accounts"],
    }
    for label, done in items.items():
        label_cell = ws.cell(row=row, column=1, value=label)
        section_style(label_cell, bg=LIGHT_GREY)
        status = "OK" if done else "MISSING"
        status_cell = ws.cell(row=row, column=2, value=status)
        value_style(status_cell, bg=(GREEN if done else RED))
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        ws.row_dimensions[row].height = 25
        row += 1

    os.makedirs(OUTPUT_PATH, exist_ok=True)
    filename = f"Handover_{data['account_name'].replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    filepath = os.path.join(OUTPUT_PATH, filename)
    wb.save(filepath)
    return filepath


def _build_from_params() -> dict:
    """Build handover data dict from consultant-provided env vars."""
    g = lambda k, d="—": os.environ.get(f"PARAM_{k.upper()}", d)
    return {
        "account_name": g("CLIENT_NAME", "Unknown Client"),
        "team": {
            "ho_done_with":   g("OUTGOING_CONSULTANT"),
            "tax_consultant": g("INCOMING_CONSULTANT"),
            "sales":          g("SALES_CONTACT"),
        },
        "company_info": {
            "sector":              g("SECTOR"),
            "type":                g("COMPANY_TYPE"),
            "account_history":     g("ACCOUNT_HISTORY"),
            "belspo_vat":          g("BELSPO_VAT"),
            "belspo_password":     g("BELSPO_PASSWORD"),
            "belspo_notification": g("BELSPO_NOTIFICATION"),
            "previous_control":    g("PREVIOUS_CONTROL"),
        },
        "contract_info": {
            "tax_measure":         g("MISSION_TYPE"),
            "expected_year":       g("MISSION_START", str(datetime.now().year)),
            "remuneration_system": g("REMUNERATION_SYSTEM"),
            "gdpr":                g("GDPR"),
        },
        "pocs": {
            "technical_contact": g("TECHNICAL_CONTACT"),
            "hr_contact":        g("HR_CONTACT"),
            "social_secretary":  g("SOCIAL_SECRETARY"),
            "key_contacts":      g("KEY_CONTACTS"),
        },
        "folder_links": {
            "sharepoint":  g("FOLDER_LINK", ""),
            "network_path": g("NETWORK_PATH", ""),
        },
        "previous_mission": {
            "technical_report":                "—",
            "timesheets":                      "—",
            "nbr_projects":                    "—",
            "nbr_eligible_employees":          "—",
            "structural_research_certificate": "—",
            "mission_status":                  "Active",
            "control":                         "—",
            "general_comments":                g("NOTES"),
            "next_step":                       g("PENDING_TASKS"),
            "introduction_email":              "—",
        },
        "belspo_notifications": [],
        "employees": [],
        "data_checklist": {
            "calculation_sheets":  False,
            "control_documents":   False,
            "diplomas":            False,
            "individual_accounts": False,
        },
    }


def run():
    log = ServiceLogger("handover-generator")
    start = time.time()
    log.info("Service started")

    filepaths = []
    try:
        # Use real consultant input if provided, otherwise fall back to mock data
        if os.environ.get("PARAM_CLIENT_NAME"):
            log.info("Running with consultant-provided parameters")
            data = _build_from_params()
            filepath = generate_handover(data)
            filepaths.append(filepath)
            log.info("Handover generated", account=data["account_name"], output_file=filepath)
        else:
            log.info("Running in demo mode with mock data")
            for data in MOCK_HANDOVER_DATA:
                filepath = generate_handover(data)
                filepaths.append(filepath)
                log.info("Handover generated", account=data["account_name"], output_file=filepath)

        duration_ms = int((time.time() - start) * 1000)
        log_run("handover-generator", status="success",
                output_file=filepaths[-1] if filepaths else None,
                duration_ms=duration_ms)
        log.info("Service completed", generated=len(filepaths), duration_ms=duration_ms)

    except Exception as exc:
        duration_ms = int((time.time() - start) * 1000)
        log.error("Service failed", error=str(exc))
        log_run("handover-generator", status="failed", error_message=str(exc), duration_ms=duration_ms)
        raise

    return filepaths


if __name__ == "__main__":
    run()
