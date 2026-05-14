import os
import sys

_SVC_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, _SVC_DIR)
for _k in list(sys.modules.keys()):
    if _k in ("mock_data", "generate_fixture", "main"):
        del sys.modules[_k]

from unittest.mock import patch
import openpyxl
import main


def test_run_succeeds_and_logs_registry():
    with patch.object(main, "log_run") as mock_log:
        result = main.run()

    assert result is not None
    assert "new" in result
    assert "processed" in result
    mock_log.assert_called()
    statuses = [c.kwargs.get("status") for c in mock_log.call_args_list]
    assert "success" in statuses


def test_output_excel_has_two_sheets():
    with patch.object(main, "log_run"):
        main.run()

    output_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "output")
    files = [f for f in os.listdir(output_dir) if f.startswith("Belspo_Notifications")]
    assert len(files) > 0

    wb = openpyxl.load_workbook(os.path.join(output_dir, files[-1]))
    assert "Notification Summary" in wb.sheetnames
    assert "Employee Details" in wb.sheetnames


def test_notifications_classified():
    with patch.object(main, "log_run"):
        result = main.run()

    total = len(result["new"]) + len(result["processed"])
    assert total > 0
