import os
import sys
import time

_SERVICES_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _SERVICES_ROOT not in sys.path:
    sys.path.insert(0, _SERVICES_ROOT)

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from mock_data import MOCK_BELSPO_NOTIFICATIONS
from shared.registry import log_run
from shared.logger import ServiceLogger

OUTPUT_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "output")

HEADER_FILL = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
NEW_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
PROCESSED_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
CENTER = Alignment(horizontal="center", vertical="center")
WRAP = Alignment(wrap_text=True, vertical="center")


def extract_notifications(notifications: list) -> dict:
    new = []
    processed = []

    for notif in notifications:
        extracted = {
            "notification_id": notif["notification_id"],
            "company": notif["company"]["name"],
            "client_id": notif["company"]["client_id"],
            "project_title": notif["project"]["title"],
            "start_date": notif["project"]["start_date"],
            "end_date": notif["project"]["end_date"],
            "research_type": notif["project"]["research_type"],
            "employee_count": len(notif["employees"]),
            "employees": notif["employees"],
            "fetched_at": notif["fetched_at"],
            "status": notif["status"],
        }

        if notif["status"] == "new":
            new.append(extracted)
        else:
            processed.append(extracted)

    return {"new": new, "processed": processed}


def export_to_excel(extracted: dict, filepath: str):
    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "Notification Summary"

    summary_headers = [
        "Notification ID", "Company", "Client ID", "Project Title",
        "Start Date", "End Date", "Research Type",
        "# Employees", "Status", "Fetched At",
    ]

    for col, header in enumerate(summary_headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER

    row = 2
    for notif in extracted["new"] + extracted["processed"]:
        values = [
            notif["notification_id"],
            notif["company"],
            notif["client_id"],
            notif["project_title"],
            notif["start_date"],
            notif["end_date"],
            notif["research_type"],
            notif["employee_count"],
            "New" if notif["status"] == "new" else "Processed",
            notif["fetched_at"],
        ]
        fill = NEW_FILL if notif["status"] == "new" else PROCESSED_FILL
        for col, value in enumerate(values, 1):
            cell = ws1.cell(row=row, column=col, value=value)
            cell.fill = fill
            cell.border = BORDER
            cell.alignment = WRAP
        row += 1

    col_widths = [25, 25, 12, 45, 12, 12, 35, 12, 15, 25]
    for i, width in enumerate(col_widths, 1):
        ws1.column_dimensions[get_column_letter(i)].width = width

    ws2 = wb.create_sheet(title="Employee Details")

    employee_headers = [
        "Notification ID", "Company", "National Registry",
        "Last Name", "First Name", "Diploma Type",
        "Diploma", "Entry Date", "Exit Date", "R&D %",
    ]

    for col, header in enumerate(employee_headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER

    row = 2
    for notif in extracted["new"] + extracted["processed"]:
        for emp in notif["employees"]:
            values = [
                notif["notification_id"],
                notif["company"],
                emp["national_registry"],
                emp["last_name"],
                emp["first_name"],
                emp["diploma_type"],
                emp["diploma"],
                emp["entry_date"],
                emp["exit_date"] or "—",
                f"{emp['rd_percentage']}%",
            ]
            fill = NEW_FILL if notif["status"] == "new" else PROCESSED_FILL
            for col, value in enumerate(values, 1):
                cell = ws2.cell(row=row, column=col, value=value)
                cell.fill = fill
                cell.border = BORDER
                cell.alignment = CENTER
            row += 1

    emp_col_widths = [25, 25, 20, 15, 15, 15, 35, 12, 12, 10]
    for i, width in enumerate(emp_col_widths, 1):
        ws2.column_dimensions[get_column_letter(i)].width = width

    wb.save(filepath)


def run():
    log = ServiceLogger("belspo-extractor")
    start = time.time()
    log.info("Service started")

    os.makedirs(OUTPUT_PATH, exist_ok=True)

    try:
        extracted = extract_notifications(MOCK_BELSPO_NOTIFICATIONS)
        log.info("Notifications extracted", new=len(extracted["new"]),
                 processed=len(extracted["processed"]))

        for notif in extracted["new"]:
            log.info("New notification", company=notif["company"],
                     project=notif["project_title"], employees=notif["employee_count"])

        filename = f"Belspo_Notifications_{datetime.now().strftime('%Y%m%d')}.xlsx"
        filepath = os.path.join(OUTPUT_PATH, filename)
        export_to_excel(extracted, filepath)

        duration_ms = int((time.time() - start) * 1000)
        log_run("belspo-extractor", status="success", output_file=filepath, duration_ms=duration_ms)
        log.info("Service completed", output_file=filepath, duration_ms=duration_ms)

    except Exception as exc:
        duration_ms = int((time.time() - start) * 1000)
        log.error("Service failed", error=str(exc))
        log_run("belspo-extractor", status="failed", error_message=str(exc), duration_ms=duration_ms)
        raise

    return extracted


if __name__ == "__main__":
    run()
