import os
import sys

_SVC_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, _SVC_DIR)
for _k in list(sys.modules.keys()):
    if _k in ("mock_data", "generate_fixture", "main"):
        del sys.modules[_k]

import openpyxl
from unittest.mock import patch, MagicMock
import main
from mock_data import MOCK_CLIENTS

MOCK_HTML = b"""
<html>
<head><title>Research and Innovation</title></head>
<body>
  <h1>Research and Innovation</h1>
  <h2>Advanced Materials Laboratory</h2>
  <p>
    We invest heavily in R&D and innovation. Our laboratories are at the forefront
    of chemistry and materials science. Numerous patents filed every year reflect
    our commitment to technological development and scientific discovery.
    Our manufacturing processes are continuously improved through applied research.
  </p>
  <h3>Digital and AI Solutions</h3>
  <p>
    We develop machine learning algorithms and data science tools to support
    our pharmaceutical and biotechnology research programs.
  </p>
</body>
</html>
"""


def _make_mock_response(status=200, content=MOCK_HTML):
    mock_resp = MagicMock()
    mock_resp.status_code = status
    mock_resp.content = content
    return mock_resp


def test_output_file_created():
    with patch.object(main, "log_run"), patch("requests.get", return_value=_make_mock_response()):
        main.run()

    output_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "output")
    files = [f for f in os.listdir(output_dir) if f.startswith("Belspo_Technical_Profile")]
    assert len(files) > 0


def test_output_has_expected_sheets():
    with patch.object(main, "log_run"), patch("requests.get", return_value=_make_mock_response()):
        main.run()

    output_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "output")
    files = sorted([f for f in os.listdir(output_dir) if f.startswith("Belspo_Technical_Profile")])
    wb = openpyxl.load_workbook(os.path.join(output_dir, files[-1]))

    assert "Company Profiles" in wb.sheetnames
    assert "Technical Report Draft" in wb.sheetnames


def test_company_profiles_has_all_clients():
    with patch.object(main, "log_run"), patch("requests.get", return_value=_make_mock_response()):
        profiles = main.run()

    assert len(profiles) == len(MOCK_CLIENTS)

    output_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "output")
    files = sorted([f for f in os.listdir(output_dir) if f.startswith("Belspo_Technical_Profile")])
    wb = openpyxl.load_workbook(os.path.join(output_dir, files[-1]))
    ws = wb["Company Profiles"]

    assert ws.cell(row=1, column=1).value == "Client Name"
    assert ws.max_row >= len(MOCK_CLIENTS) + 1


def test_rd_keywords_detected_in_mock_html():
    with patch.object(main, "log_run"), patch("requests.get", return_value=_make_mock_response()):
        profiles = main.run()

    for p in profiles:
        assert p["rd_keywords"] != "None found", f"{p['name']}: expected R&D keywords in mock HTML"


def test_failed_scrape_does_not_crash():
    with patch.object(main, "log_run"), patch("requests.get", return_value=_make_mock_response(status=503)):
        profiles = main.run()

    for p in profiles:
        assert "SCRAPING FAILED" in p["sector"]
        assert p["error"] is not None
