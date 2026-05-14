import os
import sys
import time

_SERVICES_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _SERVICES_ROOT not in sys.path:
    sys.path.insert(0, _SERVICES_ROOT)

import requests
import openpyxl
from bs4 import BeautifulSoup
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from mock_data import MOCK_CLIENTS, RD_KEYWORDS, SECTOR_KEYWORDS
from shared.registry import log_run
from shared.logger import ServiceLogger

OUTPUT_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "output")

HEADER_FILL = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
CENTER = Alignment(horizontal="center", vertical="center")
WRAP_TOP = Alignment(wrap_text=True, vertical="top")

REQUEST_TIMEOUT = 15
USER_AGENT = "Mozilla/5.0 (compatible; LeytonBelspoBot/1.0)"


def detect_sectors(text_lower):
    found = []
    for sector, keywords in SECTOR_KEYWORDS.items():
        if any(kw in text_lower for kw in keywords):
            found.append(sector)
    return found


def extract_tech_terms(soup):
    terms = []
    for tag in soup.find_all(["h1", "h2", "h3", "strong"]):
        content = tag.get_text(strip=True)
        if 4 < len(content) < 80 and content not in terms:
            terms.append(content)
    return terms[:8]


def scrape_client(client):
    name = client["name"]
    url = client["url"]

    try:
        response = requests.get(
            url,
            timeout=REQUEST_TIMEOUT,
            headers={"User-Agent": USER_AGENT},
        )
        if response.status_code != 200:
            reason = f"HTTP {response.status_code}"
            return _failed_profile(name, url, reason)

        soup = BeautifulSoup(response.content, "html.parser")

        for tag in soup(["script", "style", "nav", "footer", "noscript"]):
            tag.decompose()

        text = soup.get_text(separator=" ", strip=True)
        text_lower = text.lower()

        rd_found = [kw for kw in RD_KEYWORDS if kw in text_lower]
        sectors = detect_sectors(text_lower)
        tech_terms = extract_tech_terms(soup)
        description = " ".join(text.split())[:600]

        return {
            "name": name,
            "url": url,
            "sector": ", ".join(sectors) if sectors else "Unknown",
            "description": description,
            "rd_keywords": ", ".join(rd_found) if rd_found else "None found",
            "technologies": "; ".join(tech_terms) if tech_terms else "None found",
            "scraped_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "error": None,
        }

    except requests.exceptions.Timeout:
        return _failed_profile(name, url, "Request timed out")
    except requests.exceptions.ConnectionError as exc:
        return _failed_profile(name, url, f"Connection error: {exc}")
    except requests.exceptions.RequestException as exc:
        return _failed_profile(name, url, str(exc))


def _failed_profile(name, url, reason):
    fail_msg = f"SCRAPING FAILED — {reason}"
    return {
        "name": name,
        "url": url,
        "sector": fail_msg,
        "description": fail_msg,
        "rd_keywords": fail_msg,
        "technologies": fail_msg,
        "scraped_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "error": reason,
    }


def build_draft_text(profile):
    name = profile["name"]
    sector = profile["sector"]
    rd_kw = profile["rd_keywords"]
    tech = profile["technologies"]
    desc_excerpt = profile["description"][:400] if not profile["error"] else ""

    if profile["error"]:
        return (
            f"[DRAFT — MANUAL INPUT REQUIRED]\n\n"
            f"Automatic scraping of {name} failed: {profile['error']}.\n"
            f"Please retrieve company description and R&D activities from the client directly "
            f"or from an alternative public source.\n\n"
            f"URL attempted: {profile['url']}"
        )

    return (
        f"[DRAFT — TO BE COMPLETED BY CONSULTANT]\n\n"
        f"{name} is active in the following sector(s): {sector}.\n\n"
        f"Based on publicly available information, the company demonstrates engagement in "
        f"research and innovation activities. The following R&D-related themes were identified "
        f"on the company website: {rd_kw}.\n\n"
        f"Key technologies and competencies mentioned include: {tech}.\n\n"
        f"Public description excerpt:\n{desc_excerpt}\n\n"
        f"[Insert specific R&D project descriptions, laboratory capabilities, patent portfolio, "
        f"and innovation investment data from client documentation.]"
    )


def export_to_excel(profiles, filepath):
    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "Company Profiles"

    profile_headers = [
        "Client Name", "URL", "Sector", "Company Description",
        "R&D Keywords Found", "Technologies Mentioned", "Scraped At",
    ]
    for col, h in enumerate(profile_headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER

    for row_idx, p in enumerate(profiles, 2):
        values = [
            p["name"],
            p["url"],
            p["sector"],
            p["description"],
            p["rd_keywords"],
            p["technologies"],
            p["scraped_at"],
        ]
        for col, val in enumerate(values, 1):
            cell = ws1.cell(row=row_idx, column=col, value=val)
            cell.border = BORDER
            cell.alignment = WRAP_TOP
        ws1.row_dimensions[row_idx].height = 80

    col_widths = [22, 40, 30, 60, 40, 50, 20]
    for i, w in enumerate(col_widths, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    ws2 = wb.create_sheet(title="Technical Report Draft")

    draft_headers = ["Client Name", "Technical Report Draft"]
    for col, h in enumerate(draft_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER

    ws2.merge_cells("B1:G1")

    for row_idx, p in enumerate(profiles, 2):
        draft_text = build_draft_text(p)
        name_cell = ws2.cell(row=row_idx, column=1, value=p["name"])
        name_cell.font = Font(bold=True)
        name_cell.border = BORDER
        name_cell.alignment = Alignment(vertical="top")

        ws2.merge_cells(f"B{row_idx}:G{row_idx}")
        draft_cell = ws2.cell(row=row_idx, column=2, value=draft_text)
        draft_cell.alignment = WRAP_TOP
        draft_cell.border = BORDER
        ws2.row_dimensions[row_idx].height = 160

    ws2.column_dimensions["A"].width = 22
    for col_letter in ["B", "C", "D", "E", "F", "G"]:
        ws2.column_dimensions[col_letter].width = 20

    wb.save(filepath)


def run():
    log = ServiceLogger("web-scraper")
    start = time.time()
    log.info("Service started")

    os.makedirs(OUTPUT_PATH, exist_ok=True)

    profiles = []
    try:
        for client in MOCK_CLIENTS:
            log.info("Scraping client", client=client["name"], url=client["url"])
            profile = scrape_client(client)
            profiles.append(profile)

            if profile["error"]:
                log.warning("Scraping failed", client=client["name"], error=profile["error"])
            else:
                rd_count = len([k for k in profile["rd_keywords"].split(", ") if k != "None found"])
                log.info("Scraping succeeded", client=client["name"],
                         sector=profile["sector"], rd_keywords_found=rd_count)

        filename = f"Belspo_Technical_Profile_{datetime.now().strftime('%Y%m%d')}.xlsx"
        filepath = os.path.join(OUTPUT_PATH, filename)
        export_to_excel(profiles, filepath)

        ok_count = sum(1 for p in profiles if not p["error"])
        duration_ms = int((time.time() - start) * 1000)
        log_run("web-scraper", status="success", output_file=filepath, duration_ms=duration_ms)
        log.info("Service completed", ok=ok_count, failed=len(profiles) - ok_count,
                 output_file=filepath, duration_ms=duration_ms)

    except Exception as exc:
        duration_ms = int((time.time() - start) * 1000)
        log.error("Service failed", error=str(exc))
        log_run("web-scraper", status="failed", error_message=str(exc), duration_ms=duration_ms)
        raise

    return profiles


if __name__ == "__main__":
    run()
