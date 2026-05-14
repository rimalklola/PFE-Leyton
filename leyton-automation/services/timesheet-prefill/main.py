import os
import sys
import time

_SERVICES_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _SERVICES_ROOT not in sys.path:
    sys.path.insert(0, _SERVICES_ROOT)

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from collections import defaultdict
from mock_data import MOCK_TIMESHEETS, MOCK_PROJECTS, MONTHS
from shared.registry import log_run
from shared.logger import ServiceLogger

OUTPUT_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "output")


def get_project_rd_percentage(project_code: str) -> float:
    for project in MOCK_PROJECTS:
        if project["code"] == project_code:
            return project["rd_percentage"]
    return 0.0


def calculate_rd_hours(timesheets: list) -> list:
    results = []
    for employee_data in timesheets:
        employee = employee_data["employee"]
        rd_by_month = defaultdict(float)
        total_rd = 0.0

        for project_code, monthly_hours in employee_data["hours_per_project"].items():
            rd_pct = get_project_rd_percentage(project_code)
            for month in MONTHS:
                hours = monthly_hours.get(month, 0)
                rd_hours = hours * rd_pct
                rd_by_month[month] += rd_hours
                total_rd += rd_hours

        results.append({
            "employee": employee,
            "eligible": employee_data["eligible"],
            "diploma": employee_data["diploma"],
            "rd_hours_by_month": dict(rd_by_month),
            "total_rd_hours": round(total_rd, 2),
        })
    return results


def export_to_excel(rd_results: list, filepath: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "R&D Hours"

    header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    total_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    total_font = Font(bold=True)
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    center = Alignment(horizontal="center")

    headers = ["Employee", "Eligible", "Diploma"] + [m.capitalize() for m in MONTHS] + ["Total R&D (h)"]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    for row, emp in enumerate(rd_results, 2):
        ws.cell(row=row, column=1, value=emp["employee"]).border = border
        ws.cell(row=row, column=2, value="Oui" if emp["eligible"] else "Non").border = border
        ws.cell(row=row, column=3, value=emp["diploma"]).border = border

        for col, month in enumerate(MONTHS, 4):
            hours = emp["rd_hours_by_month"].get(month, 0)
            cell = ws.cell(row=row, column=col, value=round(hours, 2) if hours > 0 else None)
            cell.border = border
            cell.alignment = center

        total_cell = ws.cell(row=row, column=len(headers), value=emp["total_rd_hours"])
        total_cell.font = total_font
        total_cell.fill = total_fill
        total_cell.border = border
        total_cell.alignment = center

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["C"].width = 35

    wb.save(filepath)


def run():
    log = ServiceLogger("timesheet-prefill")
    start = time.time()
    log.info("Service started")

    os.makedirs(OUTPUT_PATH, exist_ok=True)

    try:
        rd_results = calculate_rd_hours(MOCK_TIMESHEETS)
        log.info("R&D hours calculated", employee_count=len(rd_results))

        filename = f"RD_Hours_{datetime.now().strftime('%Y%m%d')}.xlsx"
        filepath = os.path.join(OUTPUT_PATH, filename)
        export_to_excel(rd_results, filepath)

        for emp in rd_results:
            active = [m for m in MONTHS if emp["rd_hours_by_month"].get(m, 0) > 0]
            log.info("Employee processed", employee=emp["employee"],
                     total_rd_hours=emp["total_rd_hours"], active_months=len(active))

        duration_ms = int((time.time() - start) * 1000)
        log_run("timesheet-prefill", status="success", output_file=filepath, duration_ms=duration_ms)
        log.info("Service completed", output_file=filepath, duration_ms=duration_ms)

    except Exception as exc:
        duration_ms = int((time.time() - start) * 1000)
        log.error("Service failed", error=str(exc))
        log_run("timesheet-prefill", status="failed", error_message=str(exc), duration_ms=duration_ms)
        raise

    return rd_results


if __name__ == "__main__":
    run()
