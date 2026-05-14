import os
import sys

_SVC_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, _SVC_DIR)
for _k in list(sys.modules.keys()):
    if _k in ("mock_data", "generate_fixture", "main"):
        del sys.modules[_k]

from unittest.mock import patch
import openpyxl
import main
from mock_data import MONTHS


def test_run_succeeds_and_logs_registry():
    with patch.object(main, "log_run") as mock_log:
        results = main.run()

    assert results is not None
    assert len(results) > 0
    mock_log.assert_called()
    statuses = [c.kwargs.get("status") for c in mock_log.call_args_list]
    assert "success" in statuses


def test_total_rd_hours_nonzero():
    with patch.object(main, "log_run"):
        results = main.run()
    assert any(emp["total_rd_hours"] > 0 for emp in results)


def test_output_excel_created():
    with patch.object(main, "log_run"):
        main.run()

    output_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "output")
    files = [f for f in os.listdir(output_dir) if f.startswith("RD_Hours")]
    assert len(files) > 0

    wb = openpyxl.load_workbook(os.path.join(output_dir, files[-1]))
    ws = wb.active
    assert ws.cell(row=1, column=1).value == "Employee"
    assert ws.cell(row=1, column=len(MONTHS) + 4).value == "Total R&D (h)"
