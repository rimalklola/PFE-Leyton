import os
import sys

_SVC_DIR = os.path.dirname(os.path.abspath(__file__))
_SERVICES_ROOT = os.path.dirname(_SVC_DIR)
sys.path.insert(0, _SERVICES_ROOT)
for k in list(sys.modules.keys()):
    if k in ("main",):
        del sys.modules[k]

import openpyxl
import pytest
from unittest.mock import patch
import main


# ── Helpers ───────────────────────────────────────────────────

def make_excel_fixture(tmp_path, rows, headers=None):
    if headers is None:
        headers = ["Employee", "Month", "Hours", "RD_Percentage"]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append(row)
    path = str(tmp_path / "timesheet.xlsx")
    wb.save(path)
    return path


# ── normalize_month ───────────────────────────────────────────

def test_normalize_month_full():
    assert main.normalize_month("january") == "january"

def test_normalize_month_alias():
    assert main.normalize_month("jan") == "january"

def test_normalize_month_unknown():
    assert main.normalize_month("notamonth") is None

def test_normalize_month_case_insensitive():
    assert main.normalize_month("MARCH") == "march"


# ── extract_from_excel ────────────────────────────────────────

def test_extract_from_excel_basic(tmp_path):
    path = make_excel_fixture(tmp_path, [
        ("Alice", "january", 40, 0.9),
        ("Bob",   "february", 30, 0.8),
    ])
    rows = main.extract_from_excel(path)
    assert len(rows) == 2
    assert rows[0]["employee"] == "Alice"
    assert rows[0]["month"] == "january"
    assert rows[0]["hours"] == 40.0

def test_extract_from_excel_skips_invalid_month(tmp_path):
    path = make_excel_fixture(tmp_path, [
        ("Alice", "badmonth", 40, 0.9),
        ("Bob",   "march",    30, 0.8),
    ])
    rows = main.extract_from_excel(path)
    assert len(rows) == 1
    assert rows[0]["employee"] == "Bob"

def test_extract_from_excel_skips_empty_rows(tmp_path):
    path = make_excel_fixture(tmp_path, [
        ("Alice", "january", 40, 0.9),
        (None, None, None, None),
    ])
    rows = main.extract_from_excel(path)
    assert len(rows) == 1

def test_extract_from_excel_missing_columns_raises(tmp_path):
    path = make_excel_fixture(tmp_path, [("Alice", 40)], headers=["Employee", "Hours"])
    with pytest.raises(ValueError, match="missing required columns"):
        main.extract_from_excel(path)

def test_extract_from_excel_no_rd_percentage_defaults_to_1(tmp_path):
    path = make_excel_fixture(tmp_path, [("Alice", "march", 10)],
                              headers=["Employee", "Month", "Hours"])
    rows = main.extract_from_excel(path)
    assert rows[0]["rd_percentage"] == 1.0


# ── consolidate ───────────────────────────────────────────────

def test_consolidate_sums_same_employee_same_month():
    rows = [
        {"employee": "Alice", "month": "january", "hours": 20, "rd_percentage": 1.0},
        {"employee": "Alice", "month": "january", "hours": 10, "rd_percentage": 1.0},
    ]
    result = main.consolidate(rows)
    assert len(result) == 1
    assert result[0]["rd_hours_by_month"]["january"] == 30.0

def test_consolidate_applies_rd_percentage():
    rows = [{"employee": "Bob", "month": "february", "hours": 100, "rd_percentage": 0.7}]
    result = main.consolidate(rows)
    assert result[0]["total_rd_hours"] == 70.0

def test_consolidate_multiple_employees():
    rows = [
        {"employee": "Alice", "month": "january", "hours": 40, "rd_percentage": 1.0},
        {"employee": "Bob",   "month": "january", "hours": 30, "rd_percentage": 1.0},
    ]
    result = main.consolidate(rows)
    assert len(result) == 2


# ── export_to_excel ───────────────────────────────────────────

def test_export_creates_file(tmp_path):
    pivot = [{"employee": "Alice", "rd_hours_by_month": {"january": 36.0}, "total_rd_hours": 36.0}]
    path = str(tmp_path / "out.xlsx")
    main.export_to_excel(pivot, path)
    assert os.path.exists(path)
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    assert ws.cell(row=1, column=1).value == "Employee"
    assert ws.cell(row=2, column=1).value == "Alice"


# ── run() ─────────────────────────────────────────────────────

def test_run_with_excel_file(tmp_path):
    path = make_excel_fixture(tmp_path, [
        ("Sara Benali",   "january",  45, 0.9),
        ("Karim Idrissi", "february", 60, 0.7),
    ])
    with patch.object(main, "log_run"), \
         patch.object(main, "OUTPUT_PATH", str(tmp_path)):
        result = main.run(input_files=[path])

    assert len(result) == 2
    employees = [r["employee"] for r in result]
    assert "Sara Benali" in employees

def test_run_demo_mode_creates_fixture(tmp_path, monkeypatch):
    monkeypatch.setenv("INPUT_FILES", "")
    monkeypatch.delenv("INPUT_FILES", raising=False)
    with patch.object(main, "log_run"), \
         patch.object(main, "OUTPUT_PATH", str(tmp_path)), \
         patch.object(main, "_create_sample_fixture") as mock_fixture, \
         patch("os.path.exists", return_value=False), \
         patch.object(main, "extract_from_excel", return_value=[
             {"employee": "Demo User", "month": "january", "hours": 40.0, "rd_percentage": 1.0}
         ]):
        mock_fixture.return_value = None
        result = main.run(input_files=None)
    assert mock_fixture.called

def test_run_unsupported_file_type_skipped(tmp_path):
    fake = str(tmp_path / "file.txt")
    with open(fake, "w") as f:
        f.write("dummy")
    with patch.object(main, "log_run"), \
         patch.object(main, "OUTPUT_PATH", str(tmp_path)):
        with pytest.raises(ValueError, match="No valid timesheet data"):
            main.run(input_files=[fake])

def test_run_exception_logged(tmp_path):
    with patch.object(main, "log_run"), \
         patch.object(main, "OUTPUT_PATH", str(tmp_path)), \
         patch.object(main, "extract_from_excel", side_effect=RuntimeError("boom")):
        with pytest.raises(RuntimeError):
            main.run(input_files=[str(tmp_path / "fake.xlsx")])


# ── extract_from_pdf ─────────────────────────────────────────

def test_extract_from_pdf_basic(tmp_path):
    """Mock pdfplumber to return a well-formed table."""
    import unittest.mock as mock

    fake_table = [
        ["Employee", "Month", "Hours", "RD_Percentage"],
        ["Alice",    "january", "40",  "0.9"],
        ["Bob",      "february", "30", "0.8"],
        [None, None, None, None],
    ]

    fake_page = mock.MagicMock()
    fake_page.extract_table.return_value = fake_table
    fake_page.page_number = 1

    fake_pdf = mock.MagicMock()
    fake_pdf.__enter__ = mock.MagicMock(return_value=fake_pdf)
    fake_pdf.__exit__ = mock.MagicMock(return_value=False)
    fake_pdf.pages = [fake_page]

    with mock.patch("pdfplumber.open", return_value=fake_pdf):
        rows = main.extract_from_pdf("fake.pdf")

    assert len(rows) == 2
    assert rows[0]["employee"] == "Alice"
    assert rows[0]["month"] == "january"


def test_extract_from_pdf_no_header(tmp_path):
    """Page with no recognisable header is skipped."""
    import unittest.mock as mock

    fake_table = [["Col1", "Col2"], ["a", "b"]]
    fake_page = mock.MagicMock()
    fake_page.extract_table.return_value = fake_table
    fake_page.page_number = 1

    fake_pdf = mock.MagicMock()
    fake_pdf.__enter__ = mock.MagicMock(return_value=fake_pdf)
    fake_pdf.__exit__ = mock.MagicMock(return_value=False)
    fake_pdf.pages = [fake_page]

    with mock.patch("pdfplumber.open", return_value=fake_pdf):
        rows = main.extract_from_pdf("fake.pdf")

    assert rows == []


def test_extract_from_pdf_no_table():
    """Page with no table at all returns empty list."""
    import unittest.mock as mock

    fake_page = mock.MagicMock()
    fake_page.extract_table.return_value = None

    fake_pdf = mock.MagicMock()
    fake_pdf.__enter__ = mock.MagicMock(return_value=fake_pdf)
    fake_pdf.__exit__ = mock.MagicMock(return_value=False)
    fake_pdf.pages = [fake_page]

    with mock.patch("pdfplumber.open", return_value=fake_pdf):
        rows = main.extract_from_pdf("fake.pdf")

    assert rows == []


def test_run_with_pdf_file(tmp_path):
    """run() correctly handles a PDF input via mocked extractor."""
    import unittest.mock as mock

    fake_rows = [
        {"employee": "Sara", "month": "march", "hours": 40.0, "rd_percentage": 0.9},
    ]
    fake_pdf = str(tmp_path / "ts.pdf")
    open(fake_pdf, "w").close()

    with patch.object(main, "log_run"), \
         patch.object(main, "OUTPUT_PATH", str(tmp_path)), \
         patch.object(main, "extract_from_pdf", return_value=fake_rows):
        result = main.run(input_files=[fake_pdf])

    assert len(result) == 1
    assert result[0]["employee"] == "Sara"


# ── sample fixture creator ────────────────────────────────────

def test_create_sample_fixture(tmp_path):
    path = str(tmp_path / "sample.xlsx")
    main._create_sample_fixture(path)
    assert os.path.exists(path)
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    assert ws.cell(row=1, column=1).value == "Employee"
    assert ws.max_row > 1
