import os
import sys
import time
import json

_SERVICES_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _SERVICES_ROOT not in sys.path:
    sys.path.insert(0, _SERVICES_ROOT)

import openpyxl
import pdfplumber
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from collections import defaultdict
from shared.registry import log_run
from shared.logger import ServiceLogger

OUTPUT_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "output")

MONTHS = [
    "january", "february", "march", "april", "may", "june",
    "july", "august", "september", "october", "november", "december",
]

MONTH_ALIASES = {
    "jan": "january", "feb": "february", "mar": "march", "apr": "april",
    "jun": "june", "jul": "july", "aug": "august",
    "sep": "september", "sept": "september", "oct": "october",
    "nov": "november", "dec": "december",
}

HEADER_FILL = PatternFill(start_color="002D5B", end_color="002D5B", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TOTAL_FILL = PatternFill(start_color="E8F0FB", end_color="E8F0FB", fill_type="solid")
TOTAL_FONT = Font(bold=True, color="002D5B")
BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
CENTER = Alignment(horizontal="center", vertical="center")


def normalize_month(value: str):
    v = str(value).strip().lower()
    if v in MONTHS:
        return v
    return MONTH_ALIASES.get(v)


def extract_from_excel(filepath: str) -> list:
    """
    Reads an Excel timesheet with columns:
    Employee | Month | Hours | RD_Percentage (optional)
    Returns list of row dicts.
    """
    rows = []
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    headers = {}
    for col, cell in enumerate(ws[1], 1):
        if cell.value:
            headers[str(cell.value).strip().lower()] = col

    required = {"employee", "month", "hours"}
    if not required.issubset(set(headers.keys())):
        missing = required - set(headers.keys())
        raise ValueError(f"Excel missing required columns: {missing}")

    for row in ws.iter_rows(min_row=2, values_only=True):
        emp = row[headers["employee"] - 1]
        month = row[headers["month"] - 1]
        hours = row[headers["hours"] - 1]
        rd_pct = row[headers.get("rd_percentage", 0) - 1] if "rd_percentage" in headers else 1.0

        if not emp or not month or hours is None:
            continue

        month_norm = normalize_month(str(month))
        if not month_norm:
            continue

        try:
            rows.append({
                "employee": str(emp).strip(),
                "month": month_norm,
                "hours": float(hours),
                "rd_percentage": float(rd_pct) if rd_pct is not None else 1.0,
            })
        except (ValueError, TypeError):
            continue

    return rows


def extract_from_pdf(filepath: str) -> list:
    """
    Reads a PDF timesheet table with columns:
    Employee | Month | Hours | RD_Percentage (optional)
    """
    rows = []
    with pdfplumber.open(filepath) as pdf:
        for page in pdf.pages:
            table = page.extract_table()
            if not table:
                continue

            header_row = None
            data_start = 0
            for i, row in enumerate(table):
                if any(cell and "employee" in str(cell).lower() for cell in row):
                    header_row = [str(c).strip().lower() if c else "" for c in row]
                    data_start = i + 1
                    break

            if not header_row:
                continue

            col = {h: i for i, h in enumerate(header_row)}
            if "employee" not in col or "month" not in col or "hours" not in col:
                continue

            for row in table[data_start:]:
                if not any(row):
                    continue
                try:
                    emp = str(row[col["employee"]]).strip()
                    month_norm = normalize_month(str(row[col["month"]]))
                    hours = float(row[col["hours"]])
                    rd_pct = float(row[col["rd_percentage"]]) if "rd_percentage" in col else 1.0
                    if emp and month_norm:
                        rows.append({
                            "employee": emp,
                            "month": month_norm,
                            "hours": hours,
                            "rd_percentage": rd_pct,
                        })
                except (ValueError, TypeError):
                    continue

    return rows


def consolidate(rows: list) -> list:
    """
    Groups rows by employee and month, computes R&D hours.
    Returns list of employee dicts with hours per month and total.
    """
    by_employee = defaultdict(lambda: defaultdict(float))

    for row in rows:
        rd_hours = row["hours"] * row["rd_percentage"]
        by_employee[row["employee"]][row["month"]] += rd_hours

    result = []
    for employee, monthly in sorted(by_employee.items()):
        total = sum(monthly.values())
        result.append({
            "employee": employee,
            "rd_hours_by_month": dict(monthly),
            "total_rd_hours": round(total, 2),
        })
    return result


def export_to_excel(pivot: list, filepath: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Consolidated R&D Hours"

    headers = ["Employee"] + [m.capitalize() for m in MONTHS] + ["Total R&D (h)"]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER

    for row_idx, emp in enumerate(pivot, 2):
        ws.cell(row=row_idx, column=1, value=emp["employee"]).border = BORDER

        for col_idx, month in enumerate(MONTHS, 2):
            hours = emp["rd_hours_by_month"].get(month, 0)
            cell = ws.cell(row=row_idx, column=col_idx,
                           value=round(hours, 2) if hours > 0 else None)
            cell.border = BORDER
            cell.alignment = CENTER

        total_cell = ws.cell(row=row_idx, column=len(headers),
                             value=emp["total_rd_hours"])
        total_cell.font = TOTAL_FONT
        total_cell.fill = TOTAL_FILL
        total_cell.border = BORDER
        total_cell.alignment = CENTER

    ws.column_dimensions["A"].width = 24
    for col in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 11

    wb.save(filepath)


def run(input_files: list = None):
    """
    input_files: list of file paths (PDF or Excel) to process.
    If None, uses sample fixture for demo purposes.
    """
    log = ServiceLogger("timesheet-consolidator")
    start = time.time()
    log.info("Service started")
    os.makedirs(OUTPUT_PATH, exist_ok=True)

    try:
        # Allow params via environment variable (set by API)
        if input_files is None:
            env_files = os.environ.get("INPUT_FILES")
            if env_files:
                input_files = json.loads(env_files)
            else:
                # Demo mode: use sample fixture
                fixture = os.path.join(
                    os.path.dirname(__file__), "sample_timesheet.xlsx"
                )
                if not os.path.exists(fixture):
                    _create_sample_fixture(fixture)
                input_files = [fixture]

        all_rows = []
        for fpath in input_files:
            log.info("Processing file", path=fpath)
            ext = os.path.splitext(fpath)[1].lower()
            if ext == ".pdf":
                rows = extract_from_pdf(fpath)
            elif ext in (".xlsx", ".xls"):
                rows = extract_from_excel(fpath)
            else:
                log.warning("Unsupported file type, skipping", path=fpath)
                continue
            log.info("Rows extracted", file=os.path.basename(fpath), count=len(rows))
            all_rows.extend(rows)

        if not all_rows:
            raise ValueError("No valid timesheet data found in the provided files.")

        pivot = consolidate(all_rows)
        log.info("Consolidation complete", employee_count=len(pivot))

        filename = f"Consolidated_Timesheets_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(OUTPUT_PATH, filename)
        export_to_excel(pivot, filepath)

        duration_ms = int((time.time() - start) * 1000)
        log_run("timesheet-consolidator", status="success",
                output_file=filepath, duration_ms=duration_ms)
        log.info("Service completed", output_file=filepath,
                 employee_count=len(pivot), duration_ms=duration_ms)

    except Exception as exc:
        duration_ms = int((time.time() - start) * 1000)
        log.error("Service failed", error=str(exc))
        log_run("timesheet-consolidator", status="failed",
                error_message=str(exc), duration_ms=duration_ms)
        raise

    return pivot


def _create_sample_fixture(filepath: str):
    """Creates a sample timesheet Excel for demo mode."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Timesheet"
    ws.append(["Employee", "Month", "Hours", "RD_Percentage"])
    sample_data = [
        ("Sara Benali",    "january",  45, 0.9),
        ("Sara Benali",    "february", 50, 0.9),
        ("Sara Benali",    "march",    48, 0.9),
        ("Karim Idrissi",  "january",  60, 0.7),
        ("Karim Idrissi",  "february", 55, 0.7),
        ("Karim Idrissi",  "march",    58, 0.7),
        ("Nadia Chraibi",  "january",  40, 0.8),
        ("Nadia Chraibi",  "february", 42, 0.8),
        ("Nadia Chraibi",  "march",    38, 0.8),
    ]
    for row in sample_data:
        ws.append(row)
    wb.save(filepath)


if __name__ == "__main__":
    run()
