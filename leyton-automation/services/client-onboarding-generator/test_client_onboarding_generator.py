import os
import sys

_SVC_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, _SVC_DIR)
for _k in list(sys.modules.keys()):
    if _k in ("mock_data", "generate_fixture", "main"):
        del sys.modules[_k]

from unittest.mock import patch
import openpyxl
import main
from mock_data import MOCK_CONTRACTS


def test_run_generates_one_file_per_contract():
    with patch.object(main, "log_run") as mock_log:
        filepaths = main.run()

    assert len(filepaths) == len(MOCK_CONTRACTS)
    mock_log.assert_called()


def test_output_files_exist():
    with patch.object(main, "log_run"):
        filepaths = main.run()

    for path in filepaths:
        assert os.path.isfile(path), f"Expected file: {path}"


def test_excel_contains_expected_sections():
    with patch.object(main, "log_run"):
        filepaths = main.run()

    wb = openpyxl.load_workbook(filepaths[0])
    ws = wb.active

    cell_values = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
    sections = [v for v in cell_values if v and isinstance(v, str)]

    assert any("TIMELINE" in s for s in sections), "Expected TIMELINE section"
    assert any("DOCUMENTS" in s for s in sections), "Expected DOCUMENTS section"
    assert any("CONTACT" in s for s in sections), "Expected CONTACT section"


def test_client_name_in_filename():
    with patch.object(main, "log_run"):
        filepaths = main.run()

    for i, contract in enumerate(MOCK_CONTRACTS):
        slug = contract["client_name"].replace(" ", "_")
        assert slug in os.path.basename(filepaths[i])
