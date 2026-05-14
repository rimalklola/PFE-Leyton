import os
import sys
import time

_SERVICES_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _SERVICES_ROOT not in sys.path:
    sys.path.insert(0, _SERVICES_ROOT)

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from mock_data import MOCK_CONTRACTS, REQUIRED_DOCUMENTS, TIMELINE
from shared.registry import log_run
from shared.logger import ServiceLogger

OUTPUT_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "output")

DARK_BLUE = "003366"
LIGHT_BLUE = "D9E1F2"
WHITE = "FFFFFF"
LIGHT_GREY = "F2F2F2"

BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def _apply(cell, bg=WHITE, fg="000000", bold=False, size=11,
           h_align="left", v_align="center", wrap=False):
    cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
    cell.font = Font(color=fg, bold=bold, size=size)
    cell.alignment = Alignment(horizontal=h_align, vertical=v_align, wrap_text=wrap)
    cell.border = BORDER


def _merge_row(ws, row, col_start, col_end, value, **style_kwargs):
    cell = ws.cell(row=row, column=col_start, value=value)
    _apply(cell, **style_kwargs)
    if col_end > col_start:
        ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
    return cell


def _add_deadline(start_date_str, offset_days):
    start = datetime.strptime(start_date_str, "%Y-%m-%d")
    return (start + timedelta(days=offset_days)).strftime("%Y-%m-%d")


def generate_onboarding(contract: dict) -> str:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Onboarding Guide"

    for col, w in zip(range(1, 7), [28, 40, 18, 16, 18, 18]):
        ws.column_dimensions[get_column_letter(col)].width = w

    row = 1

    _merge_row(ws, row, 1, 6,
               f"CLIENT ONBOARDING GUIDE — {contract['client_name'].upper()}",
               bg=DARK_BLUE, fg=WHITE, bold=True, size=14, h_align="center")
    ws.row_dimensions[row].height = 40
    row += 1

    _merge_row(ws, row, 1, 6,
               f"Mission Type: {contract['mission_type']}  |  Start Date: {contract['start_date']}  "
               f"|  Consultant: {contract['consultant_name']}",
               bg=LIGHT_BLUE, fg=DARK_BLUE, bold=False, size=11, h_align="center")
    ws.row_dimensions[row].height = 25
    row += 2

    _merge_row(ws, row, 1, 6, "MISSION TIMELINE",
               bg=DARK_BLUE, fg=WHITE, bold=True, h_align="left")
    ws.row_dimensions[row].height = 22
    row += 1

    for col, h in zip(range(1, 5), ["Milestone", "Target Date", "Days from Kickoff", "Description"]):
        cell = ws.cell(row=row, column=col, value=h)
        _apply(cell, bg=LIGHT_BLUE, fg=DARK_BLUE, bold=True, h_align="center")
    ws.row_dimensions[row].height = 20
    row += 1

    for item in TIMELINE:
        target = _add_deadline(contract["start_date"], item["offset_days"])
        bg = WHITE if TIMELINE.index(item) % 2 == 0 else LIGHT_GREY
        ws.cell(row=row, column=1, value=item["milestone"])
        _apply(ws.cell(row=row, column=1), bg=bg, bold=True)
        ws.cell(row=row, column=2, value=target)
        _apply(ws.cell(row=row, column=2), bg=bg, h_align="center")
        ws.cell(row=row, column=3, value=f"J+{item['offset_days']}")
        _apply(ws.cell(row=row, column=3), bg=bg, h_align="center")
        ws.cell(row=row, column=4, value=item["description"])
        _apply(ws.cell(row=row, column=4), bg=bg, wrap=True)
        ws.row_dimensions[row].height = 22
        row += 1

    row += 1

    _merge_row(ws, row, 1, 6, "DOCUMENTS REQUIRED FROM CLIENT",
               bg=DARK_BLUE, fg=WHITE, bold=True, h_align="left")
    ws.row_dimensions[row].height = 22
    row += 1

    doc_headers = ["Document", "Description", "Format", "Deadline", "Status"]
    for col, h in enumerate(doc_headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        _apply(cell, bg=LIGHT_BLUE, fg=DARK_BLUE, bold=True, h_align="center")
    ws.row_dimensions[row].height = 20
    row += 1

    for i, doc in enumerate(REQUIRED_DOCUMENTS):
        deadline = _add_deadline(contract["start_date"], doc["deadline_days"])
        bg = WHITE if i % 2 == 0 else LIGHT_GREY
        ws.cell(row=row, column=1, value=doc["document"])
        _apply(ws.cell(row=row, column=1), bg=bg, bold=True)
        ws.cell(row=row, column=2, value=doc["description"])
        _apply(ws.cell(row=row, column=2), bg=bg, wrap=True)
        ws.cell(row=row, column=3, value=doc["format"])
        _apply(ws.cell(row=row, column=3), bg=bg, h_align="center")
        ws.cell(row=row, column=4, value=deadline)
        _apply(ws.cell(row=row, column=4), bg=bg, h_align="center")
        ws.cell(row=row, column=5, value="To Provide")
        _apply(ws.cell(row=row, column=5), bg=bg, h_align="center")
        ws.row_dimensions[row].height = 30
        row += 1

    row += 1

    _merge_row(ws, row, 1, 6, "CONTACT INFORMATION",
               bg=DARK_BLUE, fg=WHITE, bold=True, h_align="left")
    ws.row_dimensions[row].height = 22
    row += 1

    contacts = [
        ("Your Consultant", f"{contract['consultant_name']} — {contract['consultant_email']}"),
        ("Team Lead", f"{contract['tl_name']} — {contract['tl_email']}"),
        ("Leyton Morocco", "Technopark Casablanca, Route de Nouasseur, Casablanca 20100"),
    ]
    for label, value in contacts:
        ws.cell(row=row, column=1, value=label)
        _apply(ws.cell(row=row, column=1), bg=LIGHT_GREY, bold=True)
        _merge_row(ws, row, 2, 6, value, bg=WHITE)
        ws.row_dimensions[row].height = 22
        row += 1

    row += 1

    footer = (
        "Your consultant will be in touch on a quarterly basis to collect timesheets and validate R&D allocations. "
        "All documents are treated in strict confidence and processed in accordance with GDPR regulations. "
        "For any questions prior to kickoff, please reach out to your consultant directly."
    )
    _merge_row(ws, row, 1, 6, footer,
               bg=LIGHT_BLUE, fg=DARK_BLUE, wrap=True, v_align="center")
    ws.row_dimensions[row].height = 60

    client_slug = contract["client_name"].replace(" ", "_")
    filename = f"Client_Onboarding_{client_slug}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    filepath = os.path.join(OUTPUT_PATH, filename)
    os.makedirs(OUTPUT_PATH, exist_ok=True)
    wb.save(filepath)
    return filepath


def run():
    log = ServiceLogger("client-onboarding-generator")
    start = time.time()
    log.info("Service started")

    filepaths = []
    try:
        for contract in MOCK_CONTRACTS:
            filepath = generate_onboarding(contract)
            filepaths.append(filepath)
            log.info("Onboarding guide generated", client=contract["client_name"], output_file=filepath)

        duration_ms = int((time.time() - start) * 1000)
        log_run("client-onboarding-generator", status="success", duration_ms=duration_ms)
        log.info("Service completed", generated=len(filepaths), duration_ms=duration_ms)

    except Exception as exc:
        duration_ms = int((time.time() - start) * 1000)
        log.error("Service failed", error=str(exc))
        log_run("client-onboarding-generator", status="failed", error_message=str(exc), duration_ms=duration_ms)
        raise

    return filepaths


if __name__ == "__main__":
    run()
