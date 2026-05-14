import os
import sys
import time

_SERVICES_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _SERVICES_ROOT not in sys.path:
    sys.path.insert(0, _SERVICES_ROOT)

import pdfplumber
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from collections import defaultdict
from mock_data import MOCK_PROJECTS, MOCK_EMPLOYEES, MONTHS, MONTH_ALIASES
from generate_fixture import generate as generate_fixture, FIXTURE_PATH
from shared.registry import log_run
from shared.logger import ServiceLogger

OUTPUT_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "output")

COLUMN_MAP = {
    "employee name": "employee_name",
    "employee": "employee_name",
    "national registry": "national_registry",
    "registry": "national_registry",
    "project code": "project_code",
    "project": "project_code",
    "month": "month",
    "hours worked": "hours",
    "hours": "hours",
}

HEADER_FILL = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TOTAL_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
TOTAL_FONT = Font(bold=True)
BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
CENTER = Alignment(horizontal="center", vertical="center")


def normalize_month(value):
    v = str(value).strip().lower()
    if v in MONTHS:
        return v
    if v in MONTH_ALIASES:
        return MONTH_ALIASES[v]
    return None


def normalize_columns(header_row):
    mapping = {}
    for i, col in enumerate(header_row):
        if col is None:
            continue
        key = col.strip().lower()
        if key in COLUMN_MAP:
            mapping[COLUMN_MAP[key]] = i
    return mapping


def extract_rows_from_pdf(pdf_path):
    rows = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            table = page.extract_table()
            if not table:
                continue

            header_row = None
            data_start = 0
            for i, row in enumerate(table):
                if any(cell and "employee" in str(cell).lower() for cell in row):
                    header_row = row
                    data_start = i + 1
                    break

            if header_row is None:
                print(f"  WARN: No recognizable header on page {page.page_number} — skipping")
                continue

            col_map = normalize_columns(header_row)
            required = {"employee_name", "national_registry", "project_code", "month", "hours"}
            if not required.issubset(set(col_map.keys())):
                missing = required - set(col_map.keys())
                print(f"  WARN: Page {page.page_number} missing columns {missing} — skipping")
                continue

            for row in table[data_start:]:
                if not any(cell for cell in row):
                    continue
                rows.append({
                    "employee_name": row[col_map["employee_name"]],
                    "national_registry": row[col_map["national_registry"]],
                    "project_code": row[col_map["project_code"]],
                    "month": row[col_map["month"]],
                    "hours": row[col_map["hours"]],
                })
    return rows


def validate_rows(rows):
    errors = []
    for i, row in enumerate(rows, 1):
        month_norm = normalize_month(row.get("month", ""))
        if month_norm is None:
            errors.append({
                "row": i,
                "field": "month",
                "value": row.get("month"),
                "reason": "unrecognized month value",
            })

        try:
            h = float(row.get("hours", ""))
            if h < 0:
                errors.append({
                    "row": i,
                    "field": "hours",
                    "value": row.get("hours"),
                    "reason": "negative value",
                })
        except (ValueError, TypeError):
            errors.append({
                "row": i,
                "field": "hours",
                "value": row.get("hours"),
                "reason": "non-numeric value",
            })
    return errors


def save_validation_errors(errors, filepath):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Validation Errors"
    headers = ["Row", "Field", "Value", "Reason"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER
    for row_idx, err in enumerate(errors, 2):
        ws.cell(row=row_idx, column=1, value=err["row"]).border = BORDER
        ws.cell(row=row_idx, column=2, value=err["field"]).border = BORDER
        ws.cell(row=row_idx, column=3, value=str(err["value"])).border = BORDER
        ws.cell(row=row_idx, column=4, value=err["reason"]).border = BORDER
    for col in range(1, 5):
        ws.column_dimensions[get_column_letter(col)].width = 25
    wb.save(filepath)


def get_rd_percentage(project_code):
    for p in MOCK_PROJECTS:
        if p["code"] == project_code:
            return p["rd_percentage"]
    return 0.0


def get_employee_meta(name, national_registry):
    for emp in MOCK_EMPLOYEES:
        if emp["national_registry"] == national_registry or emp["name"] == name:
            return emp
    return {"name": name, "national_registry": national_registry, "eligible": True, "diploma": "Unknown"}


def compute_pivot(rows):
    by_employee = defaultdict(lambda: defaultdict(float))
    employee_meta = {}
    for row in rows:
        name = row["employee_name"]
        registry = row["national_registry"]
        project = row["project_code"]
        month = normalize_month(row["month"])
        if month is None:
            continue
        hours = float(row["hours"])
        rd_pct = get_rd_percentage(project)
        by_employee[registry][month] += hours * rd_pct
        if registry not in employee_meta:
            employee_meta[registry] = get_employee_meta(name, registry)

    results = []
    for registry, monthly in by_employee.items():
        meta = employee_meta[registry]
        total = sum(monthly.values())
        results.append({
            "employee": meta["name"],
            "eligible": meta["eligible"],
            "diploma": meta["diploma"],
            "rd_hours_by_month": dict(monthly),
            "total_rd_hours": round(total, 2),
        })
    return results


def export_to_excel(pivot, filepath):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "R&D Hours"

    headers = ["Employee", "Eligible", "Diploma"] + [m.capitalize() for m in MONTHS] + ["Total R&D (h)"]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER

    for row_idx, emp in enumerate(pivot, 2):
        ws.cell(row=row_idx, column=1, value=emp["employee"]).border = BORDER
        ws.cell(row=row_idx, column=2, value="Oui" if emp["eligible"] else "Non").border = BORDER
        ws.cell(row=row_idx, column=3, value=emp["diploma"]).border = BORDER

        for col_idx, month in enumerate(MONTHS, 4):
            hours = emp["rd_hours_by_month"].get(month, 0)
            cell = ws.cell(row=row_idx, column=col_idx, value=round(hours, 2) if hours > 0 else None)
            cell.border = BORDER
            cell.alignment = CENTER

        total_cell = ws.cell(row=row_idx, column=len(headers), value=emp["total_rd_hours"])
        total_cell.font = TOTAL_FONT
        total_cell.fill = TOTAL_FILL
        total_cell.border = BORDER
        total_cell.alignment = CENTER

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["C"].width = 32

    wb.save(filepath)


def run():
    log = ServiceLogger("pdf-timesheet-extractor")
    start = time.time()
    log.info("Service started")

    os.makedirs(OUTPUT_PATH, exist_ok=True)

    try:
        if not os.path.exists(FIXTURE_PATH):
            log.info("Generating PDF fixture")
            generate_fixture()
            log.info("Fixture generated", path=FIXTURE_PATH)

        log.info("Extracting PDF", path=FIXTURE_PATH)
        rows = extract_rows_from_pdf(FIXTURE_PATH)
        log.info("Rows extracted", count=len(rows))

        errors = validate_rows(rows)
        if errors:
            error_filename = f"Validation_Errors_{datetime.now().strftime('%Y%m%d')}.xlsx"
            error_path = os.path.join(OUTPUT_PATH, error_filename)
            save_validation_errors(errors, error_path)
            log.warning("Validation errors found", error_count=len(errors), report=error_path)
            for err in errors:
                log.warning("Validation error", row=err["row"], field=err["field"],
                            value=str(err["value"]), reason=err["reason"])
            error_rows = {e["row"] for e in errors}
            valid_rows = [r for i, r in enumerate(rows, 1) if i not in error_rows]
        else:
            valid_rows = rows
            log.info("Validation passed")

        pivot = compute_pivot(valid_rows)

        for emp in pivot:
            log.info("Employee processed", employee=emp["employee"],
                     total_rd_hours=emp["total_rd_hours"])

        filename = f"Timesheet_Extracted_{datetime.now().strftime('%Y%m%d')}.xlsx"
        filepath = os.path.join(OUTPUT_PATH, filename)
        export_to_excel(pivot, filepath)

        duration_ms = int((time.time() - start) * 1000)
        log_run("pdf-timesheet-extractor", status="success", output_file=filepath, duration_ms=duration_ms)
        log.info("Service completed", output_file=filepath, duration_ms=duration_ms)

    except Exception as exc:
        duration_ms = int((time.time() - start) * 1000)
        log.error("Service failed", error=str(exc))
        log_run("pdf-timesheet-extractor", status="failed", error_message=str(exc), duration_ms=duration_ms)
        raise

    return pivot


if __name__ == "__main__":
    run()
