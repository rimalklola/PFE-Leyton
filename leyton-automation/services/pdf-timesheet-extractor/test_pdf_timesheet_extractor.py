import os
import sys

_SVC_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, _SVC_DIR)
for _k in list(sys.modules.keys()):
    if _k in ("mock_data", "generate_fixture", "main"):
        del sys.modules[_k]

import openpyxl
from unittest.mock import patch
import main
from mock_data import MONTHS


def test_fixture_generation():
    from generate_fixture import generate, FIXTURE_PATH
    path = generate()
    assert os.path.exists(path), "Fixture PDF was not created"
    assert path.endswith(".pdf")


def test_extraction_produces_output():
    with patch.object(main, "log_run"):
        pivot = main.run()

    assert pivot is not None
    assert len(pivot) > 0

    output_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "output")
    files = sorted([f for f in os.listdir(output_dir) if f.startswith("Timesheet_Extracted")])
    assert len(files) > 0

    wb = openpyxl.load_workbook(os.path.join(output_dir, files[-1]))
    ws = wb.active
    assert ws.cell(row=1, column=1).value == "Employee"
    assert ws.cell(row=1, column=len(MONTHS) + 4).value == "Total R&D (h)"


def test_at_least_one_nonzero_rd_total():
    with patch.object(main, "log_run"):
        pivot = main.run()
    assert any(emp["total_rd_hours"] > 0 for emp in pivot)


def test_validation_passes_on_clean_fixture():
    from generate_fixture import FIXTURE_PATH, generate
    generate()
    rows = main.extract_rows_from_pdf(FIXTURE_PATH)
    errors = main.validate_rows(rows)
    assert errors == [], f"Unexpected validation errors: {errors}"
